import openpyxl as px
import re

def import_col(file):
	""" Opens an excel file and guides the user using prompts to select
	the desired cells.

	:param file: the path to the excel file
	"""

	data = [];

	wb = px.load_workbook(file)
	sheet_names = wb.get_sheet_names()
	print "Available sheets: [" + ", ".join(sheet_names) + "]"
	sheet = raw_input("Enter the name of the sheet you want to select (case-sensitive): ")
	while(sheet not in sheet_names):
		sheet = raw_input("Enter the name of the sheet you want to select (case-sensitive): ")
	ws = wb[sheet]
	col = raw_input("Enter the name of the column you want to select: ").upper()
	start = int(raw_input("Enter the first row you want to include in the selection: "))
	end = int(raw_input("Enter the last row you want to include in the selection: "))
	print "The values of your selection:"
	for i in range(start, end):
		print ws[col + str(i)].value
		data.append(ws[col + str(i)].value)

	return data


def convert_string_typed_number_to_number_typed_number(data_in):
	""" This method converts numbers of an excel-sheet which are typed as strings
	to numbers that are actually typed as numbers.
	Reasons for numbers typed as strings might be unnecessary spaces or currency
	signs.
	
	:param data_in: a string array
	"""

	data_out = []

	for i in range(0,len(data_in)):
		print "\nConverting data[" + str(i) + "]: " + str(data_in[i])
		if (data_in[i] == None): # missing value
			data_out.append(None)
			print "Converted data [" + str(i) + "] to: missing value"
		else:
			try: # let's try to convert the string
				converted_number = float(re.sub(r'\s*\$*\s*([0-9]{0,3}),*([0-9]{0,3}),*([0-9]{1,3})(\.[0-9]+)\s*', r'\1\2\3\4', str(data_in[i])))
				if (converted_number == 0): # figure out how to handle zeros
					keep_zero = raw_input("I read a zero. Do you want to keep the zero (type 0) or convert it to a missing value (leave blank)?")
					if(keep_zero == ""):
						converted_number = None
				data_out.append(converted_number)
				print "Converted data[" + str(i) + "] to: " + str(converted_number) + " (type: " + str(type(converted_number)) + ")"
			except ValueError: # couldn't convert string - ask a human for help
				not_converted = True
				print "Could not convert '" + str(data_in[i]) + "' to a number"
				while(not_converted): # for loop until we have a valid input
					manual_input = raw_input("Please enter a number or leave blank for a missing value: ")
					if (manual_input == None or manual_input == ""): # set to missing value
						data_out.append(None)
						not_converted = False
						print "Converted data[" + str(i) + "] to: missing value"
					else: # try to convert entered input
						try: 
							data_out.append(float(manual_input))
							print "Manually converted data[" + str(i) + "] to: " + manual_input
							not_converted = False
						except ValueError:
							pass

	return data_out

data = import_col("/home/sebastian/Documents/Datathon TO/MASTER - Program Tracking SheetATMarch2014.xlsx")
data = convert_string_typed_number_to_number_typed_number(data)
print "Success"