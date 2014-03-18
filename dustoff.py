import openpyxl as px
import re
import datetime
from json import dumps

def import_col(file):
	""" Opens an excel file and guides the user using prompts to select
	the desired cells. Returns a string array.

	:param file: the path to the excel file
	:returns: An object which consists of the properties 
	  * *file*: the path to the excel file
	  * *sheet*: the name of the selected sheet
	  * *col*: the name of the selected column in the sheet
	  * *start*: the first row that's included in the selected data
	  * *end*: the last row that's included in the selected data
	  * *data*: a string array containing the selected data
	"""

	data = [];

	wb = px.load_workbook(file)
	sheet_names = wb.get_sheet_names()
	print "Available sheets: [" + ", ".join(sheet_names) + "]"
	sheet = raw_input("Enter the name of the sheet you want to select (case-sensitive): ")
	while(sheet not in sheet_names):
		sheet = raw_input("Enter the name of the sheet you want to select (case-sensitive): ")
	ws = wb[sheet]
	col = raw_input("Enter the name of the column you want to select: ").upper()
	start = int(raw_input("Enter the first row you want to include in the selection: "))
	end = int(raw_input("Enter the last row you want to include in the selection: "))
	print "Data extracted from your selection:"
	for i in range(start, end + 1):
		print "Data [" + str(i - start) + "]: " + str(ws[col + str(i)].value)
		data.append(ws[col + str(i)].value)

	return { "file": file, "sheet": sheet, "col": col, "start": start, "end": end, "data": data, "wb": wb }

def convert_str_to_num(data_in):
	""" This method converts numbers of an excel-sheet which are typed as strings
	to numbers that are actually typed as numbers.
	Reasons for numbers typed as strings might be unnecessary spaces or currency
	signs.
	
	:param data_in: A string array
	:returns: An array containing the converted values typed as float. 
	"""

	data_out = []

	for i in range(0,len(data_in)):
		print "\nConverting data[" + str(i) + "]: " + str(data_in[i])
		if (data_in[i] == None): # missing value
			data_out.append(None)
			print "Converted data [" + str(i) + "] to: missing value"
		else:
			try: # let's try to convert the string
				converted_number = float(re.sub(r'\s*\$*\s*([0-9]{0,3}),*([0-9]{0,3}),*([0-9]{1,3})(\.[0-9]+)\s*', r'\1\2\3\4', str(data_in[i])))
				if (converted_number == 0): # figure out how to handle zeros
					keep_zero = raw_input("I read a zero. Do you want to keep the zero (type 0) or convert it to a missing value (leave blank)?")
					if(keep_zero == ""):
						converted_number = None
				data_out.append(converted_number)
				print "Converted data[" + str(i) + "] to: " + str(converted_number) + " (type: " + str(type(converted_number)) + ")"
			except ValueError: # couldn't convert string - ask a human for help
				not_converted = True
				print "Could not convert '" + str(data_in[i]) + "' to a number"
				while(not_converted): # for loop until we have a valid input
					manual_input = raw_input("Please enter a number or leave blank for a missing value: ")
					if (manual_input == None or manual_input == ""): # set to missing value
						data_out.append(None)
						not_converted = False
						print "Converted data[" + str(i) + "] to: missing value"
					else: # try to convert entered input
						try: 
							data_out.append(float(manual_input))
							print "Manually converted data[" + str(i) + "] to: " + manual_input
							not_converted = False
						except ValueError:
							pass

	return data_out

def clean_category(data_in):
	"""This methods cleans a column with messy categories.
	Example:
	  Say you have an array like ['class 1', 'class   1', 'clas 2', 'class 2'],
	  in this case you can use this method in order to get a clean array like
	  ['class 1', 'class 1', 'class 2', 'class 2'] which can be used for data
	  analysis.
	  
	:param data_in: A string array containing the messy data.
	:returns: A string array containing clean, categorized data.
	"""
	
	data_out = []
	categories = []
	
	next_category = raw_input("Please enter a name of a category: ")
	while(next_category != ""):
		categories.append(next_category.lower().strip())
		next_category = raw_input("So far you entered these categories: " + dumps(categories) + ". Enter another one or leave blank, when you've added all categories: ")
	
	for i in range(0, len(data_in)):
		print "\nConverting data[" + str(i) + "]: " + str(data_in[i])
		if data_in[i] is not None:
			data_in[i] = str(data_in[i]).lower().strip()
			if data_in[i] not in categories:
				not_converted = True
				print "Could not categorize " + data_in[i] + "."
				print "Available categories: "
				for j in range(0, len(categories)):
					print "[" + str(j) + "]: " + categories[j]
				while(not_converted):
					manual_input = raw_input("Please enter a category number from the list above for this cell or leave blank for a missing value: ")
					if manual_input == "":
						data_out.append(None)
						not_converted = False
						print "Manually categorized data[" + str(i) + "] to: missing value"
					elif int(manual_input) < len(categories) and int(manual_input) > -1:
						data_out.append(categories[int(manual_input)])
						not_converted = False
						print "Manually categorized data[" + str(i) + "] to: " + data_out[i]
			else:
				data_out.append(data_in[i])
				print "Categorized data[" + str(i) + "] to: " + data_out[i]
		else:
			data_out.append(None)
			print "Converted data [" + str(i) + "] to: missing value"
	
	return data_out

def save_rev(session):
	
	wb = session['wb']
	sheet = session['sheet']
	data = session['data']
	col = session['col']
	start = session['start']
	file = session['file']
	
	ws = wb[sheet]
	for i in range (0, len(data)):
		ws[col + str(start + i)] = data[i]
	
	rev = str(datetime.datetime.now()).split('.')[0].replace(":", "-")
	
	file = file.replace(".xlsx", " - rev from " + rev + ".xlsx")
	wb.save(file)
	print "Saved new revision to '" + file + "' with converted values for column '" + col + "'."